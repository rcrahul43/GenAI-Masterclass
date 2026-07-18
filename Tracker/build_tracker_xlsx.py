#!/usr/bin/env python3
"""Build GenAI-Masterclass-Tracker.xlsx with grouped sheets and cross-links."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = Path(__file__).resolve().parent / "GenAI-Masterclass-Tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUB_FONT = Font(bold=True, size=12, color="2F5496")
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="BDD7EE")
GRAY = PatternFill("solid", fgColor="E7E6E6")
YELLOW_INPUT = PatternFill("solid", fgColor="FFF2CC")

STATUS = '"Not started,In progress,Done,Interview-ready"'
YESNO = '"Yes,No"'
TRACK = '"IC,EM,Hybrid,IC+EM"'
TIER = '"Mini,Intermediate,Production,Capstone"'

SECTION_META = {
    "00-Foundations": ("00 Foundations", 0, "IC+EM"),
    "01-LLM-Engineering": ("01 LLM Engineering", 1, "IC"),
    "02-Prompt-Engineering": ("02 Prompt Engineering", 2, "IC+EM"),
    "03-Agentic-Fundamentals": ("03 Agentic Fundamentals", 3, "IC+EM"),
    "04-RAG": ("04 RAG", 4, "IC"),
    "05-Multi-Agent": ("05 Multi-Agent", 5, "IC+EM"),
    "06-Conversational-Multimodal": ("06 Voice & Multimodal", 6, "IC"),
    "07-Protocols-MCP-A2A": ("07 Protocols MCP/A2A", 7, "IC"),
    "08-Evaluation-LLMOps": ("08 Eval & LLMOps", 8, "IC+EM"),
    "09-Fine-Tuning": ("09 Fine-Tuning", 9, "IC"),
    "10-Production-Infrastructure": ("10 Production Infra", 10, "IC"),
    "11-Security-Safety": ("11 Security & Safety", 11, "IC+EM"),
    "12-Advanced-Topics": ("12 Advanced Topics", 12, "IC"),
}

WEEK_PLAN = [
    (0, "Foundations", "00-01|00-02|00-03", "BankCo decision agent", "AI judgment", 12),
    (1, "LLM Engineering I", "01-01|01-02", "Token cost estimator", "Transformer whiteboard", 12),
    (2, "LLM Engineering II", "01-03|01-04|01-05", "LiteLLM FastAPI router", "Inference tradeoffs", 12),
    (3, "Prompts & Tools", "02-01|02-02", "Structured router agent", "Tool-calling design", 12),
    (4, "Agents I", "03-01|03-02", "Inquiry routing agent", "Agent loop critique", 12),
    (5, "Agents II", "03-03|03-04", "LangGraph + checkpointing", "Pattern selection", 12),
    (6, "RAG I", "04-01|04-02", "Ingestion + chunking", "Chunking tradeoffs", 12),
    (7, "RAG II", "04-03|04-04", "Hybrid + rerank + citations", "Hallucination control", 12),
    (8, "Multi-Agent", "05-01|05-02|05-03", "Travel planner multi-agent", "Coordination failures", 14),
    (9, "Voice / Multimodal", "06-01|06-02", "ASR→LLM→TTS bot", "Latency budgets", 12),
    (10, "Protocols", "07-01|07-02|07-03", "MCP + A2A negotiation", "Protocol design", 12),
    (11, "Eval / LLMOps", "08-01|08-02|08-03", "Golden set + ship gate", "Eval strategy", 12),
    (12, "Fine-Tuning", "09-01|09-02|09-03", "LoRA compare", "Prompt vs RAG vs FT", 12),
    (13, "Production Infra", "10-01|10-02|10-03|10-04", "Dockerized agent API", "Cost/latency design", 14),
    (14, "Security + Advanced", "11-01|11-02|12-01|12-02", "Red-team + Text2SQL", "Safety review", 14),
    (15, "Advanced + Capstone", "12-03|12-04", "Capstone vertical slice", "Full design interview", 14),
    (16, "Capstone + Mocks", "CAPSTONE", "Capstone v1 + 4 mocks", "Panel simulation", 16),
]

DESIGN_LINKS = {
    "ChatGPT": "01-02, 03-01, 08-01, 11-01",
    "Claude": "01-05, 03-02, 07-01, 11-02",
    "Cursor": "03-04, 04-01, 12-03",
    "GitHub-Copilot": "01-03, 02-02, 10-04",
    "Perplexity": "04-01, 04-03, 12-01",
    "Slack-AI": "04-01, 11-01, 08-03",
    "Notion-AI": "04-01, 02-01, 08-01",
    "AI-Search-Engine": "04-03, 04-04, 10-03",
    "AI-Coding-Assistant": "03-01, 03-04, 12-03",
    "AI-Research-Agent": "12-01, 05-02, 08-01",
    "AI-Customer-Support": "03-03, 04-01, 08-03",
    "AI-Voice-Assistant": "06-01, 10-04, 08-02",
    "Multi-Agent-Workflow-Engine": "05-01, 07-01, 08-02, 10-02",
}

PROJECTS = [
    ("P01", "BankCo Decision Support", "Foundations Projects", "Mini", 0, "00-01, 00-03"),
    ("P02", "Token Cost Estimator", "LLM Projects", "Mini", 1, "01-02"),
    ("P03", "LiteLLM Gateway", "LLM Projects", "Mini", 2, "01-04, 01-05"),
    ("P04", "Structured Router Agent", "Agent Projects", "Intermediate", 3, "02-02, 03-03"),
    ("P05", "Inquiry Routing Agent", "Agent Projects", "Intermediate", 4, "03-01, 03-02"),
    ("P06", "LangGraph Production Agent", "Agent Projects", "Intermediate", 5, "03-04"),
    ("P07", "NovaCart RAG Assistant", "RAG Projects", "Production", 7, "04-01, 04-02, 04-03"),
    ("P08", "Multi-Agent Travel Planner", "Multi-Agent Projects", "Production", 8, "05-01, 05-02"),
    ("P09", "Voice Bot", "Multimodal Projects", "Intermediate", 9, "06-01"),
    ("P10", "MCP + Negotiation Sim", "Protocol Projects", "Production", 10, "07-01, 07-02, 07-03"),
    ("P11", "Eval Harness + Ship Gate", "LLMOps Projects", "Production", 11, "08-01, 08-03"),
    ("P12", "LoRA Integration", "Fine-Tune Projects", "Intermediate", 12, "09-01, 09-02, 09-03"),
    ("P13", "Deployed Agent API", "Infra Projects", "Production", 13, "10-01, 10-02, 10-04"),
    ("P14", "Capstone Multi-Agent System", "Capstone", "Capstone", 16, "05-01, 04-01, 08-01, 11-01"),
]


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, col: str, max_row: int, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max_row}")


def border_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        ws.cell(row, c).border = THIN
        ws.cell(row, c).alignment = WRAP


def first_heading(path: Path) -> str:
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if line.startswith("# "):
            return line[2:].strip()
    return path.stem


def parse_modules() -> list[dict]:
    code_to_week: dict[str, int] = {}
    for week, _theme, codes, *_rest in WEEK_PLAN:
        if codes == "CAPSTONE":
            continue
        for code in codes.split("|"):
            code_to_week[code] = week

    rows: list[dict] = []
    for path in sorted((ROOT / "Modules").rglob("*.md")):
        parts = path.parts
        section_key = parts[parts.index("Modules") + 1]
        group, order, track = SECTION_META[section_key]
        subgroup = section_key.split("-", 1)[-1].replace("-", " ")
        m = re.match(r"(\d{2}-\d{2})", path.stem)
        code = m.group(1) if m else path.stem
        title = first_heading(path)
        est, hours = "", "4"
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines()[:40]:
            if "Estimated Time" in line and "|" in line:
                parts_l = [x.strip() for x in line.split("|") if x.strip()]
                if len(parts_l) >= 2 and "Estimated" in parts_l[0]:
                    est = parts_l[1]
                    hm = re.search(r"(\d+)\s*[–-]\s*(\d+)", est) or re.search(r"(\d+)", est)
                    if hm:
                        hours = hm.group(hm.lastindex or 1)
                    break
        rows.append(
            {
                "code": code,
                "title": title,
                "group": group,
                "subgroup": subgroup,
                "section_order": order,
                "track": track,
                "hours": int(hours) if str(hours).isdigit() else hours,
                "est": est[:80],
                "file": str(path.relative_to(ROOT)),
                "week": code_to_week.get(code, order),
            }
        )
    return rows


def _split_md_table_row(line: str) -> list[str] | None:
    if not line.strip().startswith("|"):
        return None
    return [p.strip() for p in line.strip().strip("|").split("|")]


def parse_further_reading(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r"##\s+Further Reading\s*\n(.*?)(?=\n##\s|\Z)", text, re.S | re.I)
    if not m:
        return []
    block = m.group(1)
    rows: list[dict] = []
    for line in block.splitlines():
        parts = _split_md_table_row(line)
        if not parts or len(parts) < 2:
            continue
        title = parts[0]
        if title.lower() == "title" or set(title.replace(":", "")) <= {"-"}:
            continue
        url = parts[1]
        if not url.startswith("http"):
            mm = re.search(r"\((https?://[^)]+)\)", line)
            if not mm:
                continue
            url = mm.group(1)
        rows.append(
            {
                "title": title,
                "url": url,
                "difficulty": parts[2] if len(parts) > 2 else "",
                "reading_time": parts[3] if len(parts) > 3 else "",
                "why": parts[4] if len(parts) > 4 else "",
                "important": parts[5] if len(parts) > 5 else "",
            }
        )
    if not rows:
        for line in block.splitlines():
            mm = re.search(r"\[([^\]]+)\]\((https?://[^)]+)\)", line)
            if mm:
                rows.append(
                    {
                        "title": mm.group(1),
                        "url": mm.group(2),
                        "difficulty": "",
                        "reading_time": "",
                        "why": "",
                        "important": "",
                    }
                )
    return rows


def collect_resources() -> list[dict]:
    out: list[dict] = []
    for path in sorted((ROOT / "Modules").rglob("*.md")):
        section = path.parts[path.parts.index("Modules") + 1]
        m = re.match(r"(\d{2}-\d{2})", path.stem)
        item_id = m.group(1) if m else path.stem
        for i, r in enumerate(parse_further_reading(path), 1):
            out.append(
                {
                    "item_type": "Module",
                    "item_id": item_id,
                    "group": section,
                    "res_n": i,
                    "source_file": str(path.relative_to(ROOT)),
                    **r,
                }
            )
    for path in sorted((ROOT / "System Design").glob("*.md")):
        key = path.stem.replace("Design-", "")
        for i, r in enumerate(parse_further_reading(path), 1):
            out.append(
                {
                    "item_type": "System Design",
                    "item_id": key,
                    "group": "System Design",
                    "res_n": i,
                    "source_file": str(path.relative_to(ROOT)),
                    **r,
                }
            )
    for folder in ("Leadership", "Career"):
        for path in sorted((ROOT / folder).glob("*.md")):
            for i, r in enumerate(parse_further_reading(path), 1):
                out.append(
                    {
                        "item_type": folder,
                        "item_id": path.stem[:40],
                        "group": folder,
                        "res_n": i,
                        "source_file": str(path.relative_to(ROOT)),
                        **r,
                    }
                )
    return out


def build() -> Path:
    modules = parse_modules()
    resources = collect_resources()
    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "GenAI Masterclass — Progress Tracker Workbook"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "How to use",
        "1. Open Dashboard. Set G5 = your current week number.",
        "2. Change Status only on Modules, SystemDesign, LeadershipCareer, Projects, Interviews, STAR.",
        "3. Log study sessions on DailyLog (enter Module Code; Title/Week auto-fill).",
        "4. Weekly Progress % and Dashboard KPIs calculate automatically.",
        "",
        "Groups / subgroups / links",
        "Group = major section (01 LLM Engineering, 04 RAG, …)",
        "Subgroup = topic cluster inside a group",
        "Modules.Week ↔ Weekly.Module Codes (same codes)",
        "DailyLog.Module Code → VLOOKUP Title + Week from Modules",
        "Projects.Week + Linked Module Codes ↔ Weekly / Modules",
        "Interviews.Linked Modules / Linked Design ↔ study targets",
        "SystemDesign.Linked Modules ↔ Modules codes",
        "Resources sheet = Further Reading URLs extracted per Item ID (filter e.g. 04-01)",
        "Modules!Resource Count / SystemDesign!Resource Count = COUNTIFS into Resources",
        "Open Doc = opens module markdown (set Config!Link Mode + paths)",
        "Resources!URL = open reading link directly without opening the doc",
        "Resources!Open Parent Doc = jump to the chapter when you need depth",
        "",
        "Recommended workflow",
        "1) Track status on Modules / Weekly / Projects (sheet is source of truth for progress)",
        "2) Need detail? click Open Doc on that row",
        "3) Only need a paper/docs link? filter Resources by Item ID and click URL",
        "4) After editing handbook Further Reading, re-run: python Tracker/build_tracker_xlsx.py",
        "",
        "Status meanings",
        "Not started | In progress | Done | Interview-ready",
        "",
        "Open in Excel or upload to Google Sheets / Drive. See Workflow sheet.",
    ]
    for i, line in enumerate(lines, 2):
        ws[f"A{i}"] = line
        if line in {"How to use", "Groups / subgroups / links", "Status meanings"}:
            ws[f"A{i}"].font = SUB_FONT
    autosize(ws, [110])

    # Config — controls Open Doc hyperlinks
    ws = wb.create_sheet("Config")
    ws["A1"] = "Config — Doc / Resource linking"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Setting"
    ws["B3"] = "Value"
    style_header(ws, 3, 2)
    ws["A4"] = "Local Root"
    ws["B4"] = str(ROOT) + "/"
    ws["B4"].fill = YELLOW_INPUT
    ws["A5"] = "GitHub Base"
    ws["B5"] = "https://github.com/rcrahul43/GenAI-Masterclass/blob/main/"
    ws["B5"].fill = YELLOW_INPUT
    ws["A6"] = "Link Mode"
    ws["B6"] = "Local"
    ws["B6"].fill = YELLOW_INPUT
    _mode_dv = DataValidation(type="list", formula1='"Local,GitHub"', allow_blank=False)
    ws.add_data_validation(_mode_dv)
    _mode_dv.add("B6")
    ws["A8"] = "How linking works"
    ws["A8"].font = SUB_FONT
    ws["A9"] = "Local mode: Open Doc uses file:// + Local Root + File Path (works in Excel on this Mac)."
    ws["A10"] = "GitHub mode: push this folder to a repo, set GitHub Base, set Link Mode=GitHub (works in Sheets too)."
    ws["A11"] = "Resources!URL always opens the external docs/papers (no doc required)."
    ws["A12"] = "Re-sync resources from handbook: run python Tracker/build_tracker_xlsx.py"
    ws["A14"] = "Google Sheets tip"
    ws["A14"].font = SUB_FONT
    ws["A15"] = "Upload xlsx to Drive → Open with Google Sheets. Use Link Mode=GitHub for clickable docs."
    ws["A16"] = "Keep progress columns (Status, minutes, Done?) only in the Sheet — don’t duplicate in markdown."
    autosize(ws, [18, 88])

    # Workflow
    ws = wb.create_sheet("Workflow")
    ws["A1"] = "Sheet ↔ Doc ↔ Resource workflow"
    ws["A1"].font = TITLE_FONT
    flow = [
        "",
        "Goal",
        "Sheet = progress dashboard. Doc = deep reading when needed. Resource URL = quick official link.",
        "",
        "Path A — study with depth",
        "Dashboard / Modules → set Status=In progress → click Open Doc → read chapter → open Further Reading from doc OR from Resources sheet → mark Status=Done",
        "",
        "Path B — no doc needed",
        "Modules → note Item ID → Resources filter Item ID → click URL → mark Resources Done?=Yes → optionally mark module Done if that was enough",
        "",
        "What syncs automatically",
        "Handbook Further Reading → Resources sheet (when you re-run build_tracker_xlsx.py)",
        "Modules Resource Count ← formula into Resources",
        "DailyLog Title/Week ← VLOOKUP Modules",
        "Weekly Progress % ← formulas from Modules Status",
        "",
        "What does NOT auto-sync (by design)",
        "Your Status / minutes / mock scores stay in the Sheet only (source of truth for progress)",
        "Markdown checkboxes are optional backup, not required",
        "",
        "Keep in sync after editing docs",
        "1. Edit module Further Reading in markdown",
        "2. cd GenAI-Masterclass && python3 Tracker/build_tracker_xlsx.py",
        "3. Re-open xlsx (or re-upload to Sheets). Progress columns you typed are rebuilt — export Status first if needed",
        "",
        "Safer sync habit",
        "Treat Sheet Status as sacred: before rebuild, copy Modules!A:I to a backup tab, rebuild, then paste Status back by Module Code.",
    ]
    for i, line in enumerate(flow, 2):
        ws[f"A{i}"] = line
        if line in {
            "Goal",
            "Path A — study with depth",
            "Path B — no doc needed",
            "What syncs automatically",
            "What does NOT auto-sync (by design)",
            "Keep in sync after editing docs",
            "Safer sync habit",
        }:
            ws[f"A{i}"].font = SUB_FONT
    autosize(ws, [120])

    # Lookups
    ws = wb.create_sheet("Lookups")
    headers = ["Status", "Track", "Tier", "MockType", "YesNo", "STAR Pattern"]
    ws.append(headers)
    style_header(ws, 1, 6)
    for i, s in enumerate(["Not started", "In progress", "Done", "Interview-ready"], 2):
        ws[f"A{i}"] = s
    for i, s in enumerate(["IC", "EM", "Hybrid", "IC+EM"], 2):
        ws[f"B{i}"] = s
    for i, s in enumerate(["Mini", "Intermediate", "Production", "Capstone"], 2):
        ws[f"C{i}"] = s
    for i, s in enumerate(
        [
            "Senior AI",
            "Staff AI",
            "Principal Design",
            "EM Behavioral",
            "Staff+Security",
            "EM Execution",
            "Full Panel",
            "Offer Rehearsal",
        ],
        2,
    ):
        ws[f"D{i}"] = s
    ws["E2"] = "Yes"
    ws["E3"] = "No"
    for i, s in enumerate(
        [
            "Project Experience",
            "Conflict Resolution",
            "Execution",
            "Hiring",
            "Technical Judgment",
            "Cross-Functional",
            "Leadership Profile",
            "Personal Growth",
        ],
        2,
    ):
        ws[f"F{i}"] = s
    autosize(ws, [18, 12, 14, 18, 10, 22])

    # Modules
    ws = wb.create_sheet("Modules")
    headers = [
        "Module Code",
        "Title",
        "Group",
        "Subgroup",
        "Section Order",
        "Week",
        "Track",
        "Est Hours",
        "Status",
        "Minutes Spent",
        "Lab Done",
        "Judgment Memo",
        "Interview Drill",
        "Next Action",
        "File Path",
        "Open Doc",
        "Notes",
        "Resource Count",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    open_doc_formula = (
        '=IF(O{r}="","",'
        'IF(Config!$B$6="GitHub",'
        'HYPERLINK(Config!$B$5&O{r},"Open doc"),'
        'HYPERLINK("file://"&Config!$B$4&O{r},"Open doc")))'
    )
    for m in modules:
        r = ws.max_row + 1
        ws.append(
            [
                m["code"],
                m["title"],
                m["group"],
                m["subgroup"],
                m["section_order"],
                m["week"],
                m["track"],
                m["hours"],
                "Not started",
                "",
                "No",
                "No",
                "No",
                "",
                m["file"],
                open_doc_formula.format(r=r),
                "",
                f'=COUNTIFS(Resources!A:A,"Module",Resources!B:B,A{r})',
            ]
        )
        border_row(ws, ws.max_row, len(headers))
        ws.cell(r, 16).font = LINK_FONT
    last = ws.max_row
    add_dv(ws, "I", last, STATUS)
    for col in ("K", "L", "M"):
        add_dv(ws, col, last, YESNO)
    ws.conditional_formatting.add(f"I2:I{last}", FormulaRule(formula=['$I2="Done"'], fill=GREEN))
    ws.conditional_formatting.add(f"I2:I{last}", FormulaRule(formula=['$I2="In progress"'], fill=YELLOW))
    ws.conditional_formatting.add(f"I2:I{last}", FormulaRule(formula=['$I2="Interview-ready"'], fill=BLUE))
    ws.conditional_formatting.add(f"I2:I{last}", FormulaRule(formula=['$I2="Not started"'], fill=GRAY))
    ws.auto_filter.ref = f"A1:R{last}"
    ws.freeze_panes = "B2"
    autosize(ws, [12, 55, 22, 18, 8, 8, 10, 10, 16, 12, 10, 12, 12, 28, 45, 12, 24, 14])

    # Resources (Further Reading extracted from chapters)
    ws = wb.create_sheet("Resources")
    headers = [
        "Item Type",
        "Item ID",
        "Group",
        "Res #",
        "Title",
        "URL",
        "Open Resource",
        "Open Parent Doc",
        "Difficulty",
        "Reading Time",
        "Why Read",
        "Important Sections",
        "Source File",
        "Done?",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    parent_doc_formula = (
        '=IF(M{r}="","",'
        'IF(Config!$B$6="GitHub",'
        'HYPERLINK(Config!$B$5&M{r},"Open parent doc"),'
        'HYPERLINK("file://"&Config!$B$4&M{r},"Open parent doc")))'
    )
    for res in resources:
        r = ws.max_row + 1
        ws.append(
            [
                res["item_type"],
                res["item_id"],
                res["group"],
                res["res_n"],
                res["title"],
                res["url"],
                f'=IF(F{r}="","",HYPERLINK(F{r},"Open link"))',
                parent_doc_formula.format(r=r),
                res["difficulty"],
                res["reading_time"],
                res["why"],
                res["important"],
                res["source_file"],
                "No",
                "",
            ]
        )
        border_row(ws, ws.max_row, len(headers))
        ws.cell(r, 6).hyperlink = res["url"]
        ws.cell(r, 6).font = LINK_FONT
        ws.cell(r, 7).font = LINK_FONT
        ws.cell(r, 8).font = LINK_FONT
    last = ws.max_row
    add_dv(ws, "N", last, YESNO)
    ws.auto_filter.ref = f"A1:O{last}"
    ws.freeze_panes = "E2"
    autosize(ws, [14, 22, 22, 6, 42, 45, 12, 16, 14, 12, 36, 32, 40, 8, 20])

    # SystemDesign
    ws = wb.create_sheet("SystemDesign")
    headers = [
        "Design ID",
        "Name",
        "Group",
        "Subgroup",
        "Track",
        "Est Hours",
        "Status",
        "Week Target",
        "Linked Modules",
        "Minutes Spent",
        "Writeup Done",
        "Mock Used",
        "File Path",
        "Open Doc",
        "Notes",
        "Resource Count",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    open_design_formula = (
        '=IF(M{r}="","",'
        'IF(Config!$B$6="GitHub",'
        'HYPERLINK(Config!$B$5&M{r},"Open doc"),'
        'HYPERLINK("file://"&Config!$B$4&M{r},"Open doc")))'
    )
    for path in sorted((ROOT / "System Design").glob("*.md")):
        key = path.stem.replace("Design-", "")
        r = ws.max_row + 1
        rel = str(path.relative_to(ROOT))
        ws.append(
            [
                key,
                first_heading(path),
                "System Design",
                "Product Design",
                "IC+EM",
                3,
                "Not started",
                "",
                DESIGN_LINKS.get(key, ""),
                "",
                "No",
                "No",
                rel,
                open_design_formula.format(r=r),
                "",
                f'=COUNTIFS(Resources!A:A,"System Design",Resources!B:B,A{r})',
            ]
        )
        border_row(ws, ws.max_row, len(headers))
        ws.cell(r, 14).font = LINK_FONT
    last = ws.max_row
    add_dv(ws, "G", last, STATUS)
    add_dv(ws, "K", last, YESNO)
    add_dv(ws, "L", last, YESNO)
    ws.conditional_formatting.add(f"G2:G{last}", FormulaRule(formula=['$G2="Done"'], fill=GREEN))
    ws.conditional_formatting.add(f"G2:G{last}", FormulaRule(formula=['$G2="In progress"'], fill=YELLOW))
    ws.auto_filter.ref = f"A1:P{last}"
    ws.freeze_panes = "B2"
    autosize(ws, [24, 42, 14, 14, 10, 10, 14, 12, 32, 12, 12, 10, 42, 12, 20, 14])

    # LeadershipCareer
    ws = wb.create_sheet("LeadershipCareer")
    headers = [
        "Item ID",
        "Name",
        "Group",
        "Subgroup",
        "Track",
        "Est Hours",
        "Status",
        "Linked Modules",
        "STAR Related",
        "File Path",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for folder, subgroup, track in (
        ("Leadership", "Leadership", "EM"),
        ("Career", "Career", "IC+EM"),
    ):
        for path in sorted((ROOT / folder).glob("*.md")):
            item_id = path.stem[:32]
            star = "Yes" if any(k in item_id for k in ("STAR", "Behavioral", "EM-Interview", "Principal")) else "No"
            ws.append(
                [
                    item_id,
                    first_heading(path),
                    "Leadership & Career",
                    subgroup,
                    track,
                    3,
                    "Not started",
                    "00-01, 08-03, 11-01",
                    star,
                    str(path.relative_to(ROOT)),
                    "",
                ]
            )
            border_row(ws, ws.max_row, len(headers))
    last = ws.max_row
    add_dv(ws, "G", last, STATUS)
    ws.auto_filter.ref = f"A1:K{last}"
    ws.freeze_panes = "B2"
    autosize(ws, [30, 48, 20, 12, 10, 10, 14, 24, 12, 48, 20])

    # Weekly
    ws = wb.create_sheet("Weekly")
    headers = [
        "Week",
        "Theme",
        "Group",
        "Module Codes",
        "Project",
        "Interview Focus",
        "Hours Planned",
        "Hours Actual",
        "Status",
        "Modules Total",
        "Modules Done+IR",
        "Week Progress %",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for week, theme, codes, proj, iv, hrs in WEEK_PLAN:
        if codes == "CAPSTONE":
            group = "Capstone"
            code_list: list[str] = []
        else:
            code_list = codes.split("|")
            prefix = code_list[0][:2]
            group = next((v[0] for k, v in SECTION_META.items() if k.startswith(prefix)), "")
        r = ws.max_row + 1
        ws.append(
            [
                week,
                theme,
                group,
                ", ".join(code_list) if code_list else "CAPSTONE",
                proj,
                iv,
                hrs,
                "",
                "Not started",
                len(code_list) if code_list else 1,
            ]
        )
        if code_list:
            parts = [
                f'COUNTIFS(Modules!$A:$A,"{c}",Modules!$I:$I,"Done")+COUNTIFS(Modules!$A:$A,"{c}",Modules!$I:$I,"Interview-ready")'
                for c in code_list
            ]
            ws.cell(r, 11, f'={"+".join(parts)}')
        else:
            ws.cell(
                r,
                11,
                '=COUNTIFS(Projects!D:D,"Capstone",Projects!F:F,"Done")+COUNTIFS(Projects!D:D,"Capstone",Projects!F:F,"Interview-ready")',
            )
        ws.cell(r, 12, f"=IF(J{r}=0,\"\",K{r}/J{r})")
        ws.cell(r, 12).number_format = "0%"
        border_row(ws, r, len(headers))
    last = ws.max_row
    add_dv(ws, "I", last, STATUS)
    ws.auto_filter.ref = f"A1:M{last}"
    ws.freeze_panes = "C2"
    autosize(ws, [8, 22, 20, 34, 28, 26, 12, 12, 14, 12, 14, 12, 24])

    # DailyLog
    ws = wb.create_sheet("DailyLog")
    headers = [
        "Date",
        "Week",
        "Module Code",
        "Module Title (auto)",
        "Track Day",
        "Concept min",
        "Build min",
        "Judgment min",
        "Interview min",
        "Total min",
        "Status",
        "Blocker",
        "Win",
        "Next Action",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r in range(2, 52):
        ws.cell(r, 2, f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},Modules!$A:$F,6,FALSE),""))')
        ws.cell(r, 4, f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},Modules!$A:$B,2,FALSE),""))')
        ws.cell(r, 10, f"=SUM(F{r}:I{r})")
        ws.cell(r, 11, "Not started")
        border_row(ws, r, len(headers))
    add_dv(ws, "K", 51, STATUS)
    add_dv(ws, "E", 51, '"IC,EM,Hybrid,Rest"')
    ws.auto_filter.ref = "A1:N51"
    ws.freeze_panes = "D2"
    autosize(ws, [12, 8, 12, 48, 10, 12, 10, 12, 12, 10, 14, 24, 24, 24])

    # Projects
    ws = wb.create_sheet("Projects")
    headers = [
        "Project ID",
        "Name",
        "Group",
        "Tier",
        "Week",
        "Status",
        "Linked Module Codes",
        "Repo URL",
        "Resume Bullet Ready",
        "Demo Ready",
        "Architecture Review",
        "Code Review",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in PROJECTS:
        ws.append([row[0], row[1], row[2], row[3], row[4], "Not started", row[5], "", "No", "No", "No", "No", ""])
        border_row(ws, ws.max_row, len(headers))
    last = ws.max_row
    add_dv(ws, "F", last, STATUS)
    add_dv(ws, "D", last, TIER)
    for col in ("I", "J", "K", "L"):
        add_dv(ws, col, last, YESNO)
    ws.conditional_formatting.add(f"F2:F{last}", FormulaRule(formula=['$F2="Done"'], fill=GREEN))
    ws.conditional_formatting.add(f"F2:F{last}", FormulaRule(formula=['$F2="In progress"'], fill=YELLOW))
    ws.auto_filter.ref = f"A1:M{last}"
    ws.freeze_panes = "B2"
    autosize(ws, [10, 32, 20, 14, 8, 14, 30, 28, 14, 12, 14, 12, 24])

    # Interviews
    ws = wb.create_sheet("Interviews")
    headers = [
        "Mock #",
        "Date",
        "Type",
        "Role Target",
        "Score Avg",
        "Weak Dimension",
        "Linked Modules",
        "Linked Design",
        "What Went Well",
        "Critical Miss",
        "Retest Date",
        "Status",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    mocks = [
        (1, "Senior AI", "IC-AG", "03-01, 03-02", ""),
        (2, "Staff AI", "IC-AR", "04-01, 04-03", ""),
        (3, "Principal Design", "IC-COST", "05-01, 10-04", "Cursor"),
        (4, "EM Behavioral", "EM-PEO", "", ""),
        (5, "Staff+Security", "IC-SEC", "11-01, 11-02", ""),
        (6, "EM Execution", "EM-EXE", "", ""),
        (7, "Full Panel", "IC-COM", "03-04, 08-01", "ChatGPT"),
        (8, "Offer Rehearsal", "EM-ROI", "", ""),
    ]
    for mock in mocks:
        ws.append([mock[0], "", mock[1], "", "", mock[2], mock[3], mock[4], "", "", "", "Planned", ""])
        border_row(ws, ws.max_row, len(headers))
    last = ws.max_row
    dv = DataValidation(type="list", formula1='"Planned,Done,Needs retest"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{last}")
    ws.auto_filter.ref = f"A1:M{last}"
    autosize(ws, [8, 12, 18, 14, 10, 14, 22, 14, 28, 28, 12, 14, 20])

    # STAR
    ws = wb.create_sheet("STAR")
    headers = [
        "Story ID",
        "Title",
        "Pattern",
        "Group",
        "Role Angle",
        "Status",
        "Strength 1-5",
        "Linked Modules",
        "Last Practiced",
        "Situation",
        "Task",
        "Action",
        "Result",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    stories = [
        ("S01", "Led AI feature idea→launch", "Project Experience", "Delivery", "Both", "00-01, 08-03"),
        ("S02", "Cross-team alignment conflict", "Conflict Resolution", "People", "EM", "05-01"),
        ("S03", "Missed deadline recovery", "Execution", "Delivery", "Both", "10-02"),
        ("S04", "Hired/leveled AI engineer", "Hiring", "People", "EM", "08-01"),
        ("S05", "Hallucination/prod incident", "Technical Judgment", "Tech", "IC", "04-01, 11-02"),
        ("S06", "Said no to fine-tune/multi-agent", "Technical Judgment", "Tech", "IC", "09-02, 05-01"),
        ("S07", "Influenced without authority", "Cross-Functional", "People", "Both", "00-01"),
        ("S08", "Built mentorship rituals", "Leadership Profile", "People", "EM", "08-03"),
    ]
    for story in stories:
        ws.append([story[0], story[1], story[2], story[3], story[4], "Draft", "", story[5], "", "", "", "", "", ""])
        border_row(ws, ws.max_row, len(headers))
    last = ws.max_row
    dv = DataValidation(type="list", formula1='"Draft,Ready,Interview-polished"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{last}")
    ws.auto_filter.ref = f"A1:N{last}"
    autosize(ws, [10, 36, 20, 12, 10, 16, 12, 20, 12, 24, 24, 24, 24, 16])

    # Helper for next modules
    helper = wb.create_sheet("_Helper")
    helper.append(["Code", "Title", "Group", "Week", "Status", "Include", "Rank"])
    for i, m in enumerate(modules, 2):
        helper.cell(i, 1, m["code"])
        helper.cell(i, 2, m["title"])
        helper.cell(i, 3, m["group"])
        helper.cell(i, 4, m["week"])
        helper.cell(i, 5, f'=IFERROR(VLOOKUP(A{i},Modules!$A:$I,9,FALSE),"")')
        helper.cell(i, 6, f'=IF(OR(E{i}="Not started",E{i}="In progress"),1,0)')
        helper.cell(i, 7, f'=IF(F{i}=0,"",COUNTIFS($F$2:F{i},1))')
    helper.sheet_state = "hidden"

    # Dashboard
    ws = wb.create_sheet("Dashboard", 1)
    ws["A1"] = "GenAI Masterclass Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = "Yellow cell G5 = current week. Update Status on other sheets; this page calculates."
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A3"] = "Track"
    ws["B3"] = "IC"
    ws["B3"].fill = YELLOW_INPUT
    add_dv(ws, "B", 3, TRACK)

    ws["A5"] = "KPIs"
    ws["A5"].font = SUB_FONT
    kpis = [
        (6, "Modules Done+IR", '=COUNTIF(Modules!I:I,"Done")+COUNTIF(Modules!I:I,"Interview-ready")'),
        (7, "Modules Total", str(len(modules))),
        (8, "Module Progress %", "=IF(B7=0,0,B6/B7)"),
        (9, "Projects Done+IR", '=COUNTIF(Projects!F:F,"Done")+COUNTIF(Projects!F:F,"Interview-ready")'),
        (10, "Projects Total", str(len(PROJECTS))),
        (11, "Mocks Done", '=COUNTIF(Interviews!L:L,"Done")'),
        (12, "STAR Ready+", '=COUNTIF(STAR!F:F,"Ready")+COUNTIF(STAR!F:F,"Interview-polished")'),
        (13, "Hours Logged (min)", "=SUM(DailyLog!J:J)"),
    ]
    for row, label, formula in kpis:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, formula)
        ws.cell(row, 2).border = THIN
    ws["B8"].number_format = "0%"

    ws["A15"] = "Progress by Group"
    ws["A15"].font = SUB_FONT
    ws["A16"] = "Group"
    ws["B16"] = "Total"
    ws["C16"] = "Done+IR"
    ws["D16"] = "Progress %"
    style_header(ws, 16, 4)
    groups = []
    seen = set()
    for m in modules:
        if m["group"] not in seen:
            seen.add(m["group"])
            groups.append(m["group"])
    for i, g in enumerate(groups):
        r = 17 + i
        ws.cell(r, 1, g)
        ws.cell(r, 2, f"=COUNTIF(Modules!C:C,A{r})")
        ws.cell(
            r,
            3,
            f'=COUNTIFS(Modules!C:C,A{r},Modules!I:I,"Done")+COUNTIFS(Modules!C:C,A{r},Modules!I:I,"Interview-ready")',
        )
        ws.cell(r, 4, f"=IF(B{r}=0,0,C{r}/B{r})")
        ws.cell(r, 4).number_format = "0%"
        for c in range(1, 5):
            ws.cell(r, c).border = THIN
    group_end = 16 + len(groups)

    ws["F5"] = "This Week Focus"
    ws["F5"].font = SUB_FONT
    ws["F6"] = "Week # →"
    ws["G6"] = 0
    ws["G6"].fill = YELLOW_INPUT
    ws["G6"].border = THIN
    ws["F7"] = "Theme"
    ws["G7"] = '=IFERROR(VLOOKUP($G$6,Weekly!A:B,2,FALSE),"")'
    ws["F8"] = "Modules"
    ws["G8"] = '=IFERROR(VLOOKUP($G$6,Weekly!A:D,4,FALSE),"")'
    ws["F9"] = "Project"
    ws["G9"] = '=IFERROR(VLOOKUP($G$6,Weekly!A:E,5,FALSE),"")'
    ws["F10"] = "Interview focus"
    ws["G10"] = '=IFERROR(VLOOKUP($G$6,Weekly!A:F,6,FALSE),"")'
    ws["F11"] = "Week Progress %"
    ws["G11"] = '=IFERROR(VLOOKUP($G$6,Weekly!A:L,12,FALSE),"")'
    ws["G11"].number_format = "0%"
    for r in range(7, 12):
        ws.cell(r, 6).font = Font(bold=True)
        ws.cell(r, 7).border = THIN

    ws["F13"] = "Next 5 Modules"
    ws["F13"].font = SUB_FONT
    ws["F14"] = "Code"
    ws["G14"] = "Title"
    ws["H14"] = "Group"
    ws["I14"] = "Week"
    style_header(ws, 14, 4)
    # fix header style overwrote texts
    ws["F14"] = "Code"
    ws["G14"] = "Title"
    ws["H14"] = "Group"
    ws["I14"] = "Week"
    style_header(ws, 14, 4)
    for i in range(5):
        r = 15 + i
        n = i + 1
        ws.cell(r, 6, f"=IFERROR(INDEX(_Helper!A:A,MATCH({n},_Helper!G:G,0)),\"\")")
        ws.cell(r, 7, f"=IFERROR(INDEX(_Helper!B:B,MATCH({n},_Helper!G:G,0)),\"\")")
        ws.cell(r, 8, f"=IFERROR(INDEX(_Helper!C:C,MATCH({n},_Helper!G:G,0)),\"\")")
        ws.cell(r, 9, f"=IFERROR(INDEX(_Helper!D:D,MATCH({n},_Helper!G:G,0)),\"\")")
        for c in range(6, 10):
            ws.cell(r, c).border = THIN

    tier_start = group_end + 2
    ws.cell(tier_start, 1, "Projects by Tier").font = SUB_FONT
    ws.cell(tier_start + 1, 1, "Tier")
    ws.cell(tier_start + 1, 2, "Total")
    ws.cell(tier_start + 1, 3, "Done+IR")
    ws.cell(tier_start + 1, 4, "Progress %")
    style_header(ws, tier_start + 1, 4)
    for i, tier in enumerate(["Mini", "Intermediate", "Production", "Capstone"]):
        r = tier_start + 2 + i
        ws.cell(r, 1, tier)
        ws.cell(r, 2, f"=COUNTIF(Projects!D:D,A{r})")
        ws.cell(
            r,
            3,
            f'=COUNTIFS(Projects!D:D,A{r},Projects!F:F,"Done")+COUNTIFS(Projects!D:D,A{r},Projects!F:F,"Interview-ready")',
        )
        ws.cell(r, 4, f"=IF(B{r}=0,0,C{r}/B{r})")
        ws.cell(r, 4).number_format = "0%"
        for c in range(1, 5):
            ws.cell(r, c).border = THIN

    ir = tier_start + 7
    ws.cell(ir, 1, "Interview Pipeline").font = SUB_FONT
    ws.cell(ir + 1, 1, "Planned")
    ws.cell(ir + 1, 2, '=COUNTIF(Interviews!L:L,"Planned")')
    ws.cell(ir + 2, 1, "Done")
    ws.cell(ir + 2, 2, '=COUNTIF(Interviews!L:L,"Done")')
    ws.cell(ir + 3, 1, "Needs retest")
    ws.cell(ir + 3, 2, '=COUNTIF(Interviews!L:L,"Needs retest")')
    ws.cell(ir + 4, 1, "Avg Score (Done)")
    ws.cell(ir + 4, 2, '=IFERROR(AVERAGEIF(Interviews!L:L,"Done",Interviews!E:E),"")')

    ws["F21"] = "Jump to sheet"
    ws["F21"].font = SUB_FONT
    for i, name in enumerate(
        [
            "Workflow",
            "Config",
            "Modules",
            "Resources",
            "Weekly",
            "DailyLog",
            "Projects",
            "Interviews",
            "STAR",
            "SystemDesign",
            "LeadershipCareer",
        ]
    ):
        cell = ws.cell(22 + i, 6, name)
        cell.hyperlink = f"#'{name}'!A1"
        cell.font = LINK_FONT

    ws["F31"] = "Tip: On Modules sheet use the filter arrows on Group / Subgroup / Status / Week."
    ws["F32"] = "Tip: Filter Projects by Tier for a board-like pipeline."
    autosize(ws, [28, 14, 12, 12, 8, 18, 42, 18, 10])

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    out = build()
    print(f"Wrote {out}")
